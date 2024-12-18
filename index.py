import pandas as pd
import json

from openpyxl           import load_workbook
from openpyxl.styles    import Alignment
from openpyxl.styles    import PatternFill


from function.movimentacoes         import cont_movimentacoes
from function.beneficios            import cont_beneficio
from function.trabalho              import cont_trabalho
from function.estrangeiros          import cont_estrangeiros
from function.juridico              import cont_juridico
from function.gerarTaxaDeOcupacao   import criar_tx_ocupacao
from function.preencherTxOcupacao   import preencher_tx_ocupacao
from function.gerarDevolutiva       import gerar_devolutiva
from function.gerarMrosc            import gerar_mrosc
from function.preencherDevolutiva   import preencher_devolutiva
from function.preencherMROSC        import preencher_mrosc


try:
    with open("dados.json", "r") as f:
        dados = json.load(f)  # Carrega o JSON completo em um dicionário
        mes_ref = dados["mes_ref"]  # Lê a chave 'mes_ref'
        ini_acolhidos_valor = dados["ini_acolhidos"]  # Lê a chave 'ini_acolhidos'
        
except FileNotFoundError:
    print("Arquivo dados.json não encontrado. Certifique-se de rodar o segundo programa primeiro.")
    mes_ref = None

nome_tabela_ppa = "PLANILHA DE MONITORAMENTO - ALTA COMPLEXIDADE - EIXO ADULTO - ALBERGUE MARTIN LUTHER KING JR.xlsx"

df = pd.read_excel(nome_tabela_ppa, skiprows=5)

#CONVERSÃO DAS COLUNAS PARA DATETIME
df['DATA DO ACOLHIMENTO ATUAL DD/MM/AAAA']  = pd.to_datetime(df['DATA DO ACOLHIMENTO ATUAL DD/MM/AAAA'], errors='coerce')
df['DATA DO DESLIGAMENTO DD/MM/AAAA']       = pd.to_datetime(df['DATA DO DESLIGAMENTO DD/MM/AAAA'], errors='coerce')


#mes_ref                 = '11_NOV_2024'
nome_tb_tx_ocupacao     = 'TaxaDeOcupacao.xlsx'
nome_tb_devolutiva      = 'DEVOLUTIVA.xlsx'
nome_tb_mrosc           = "MROSC.xlsx"
mes                     = int(mes_ref[:2])
ano                     = int(mes_ref[-4:])
ini_acolhidos           = ini_acolhidos_valor


#ONTAR MOVIMENTAÇÕES DE USUÁRIOS
movimentacoes   = cont_movimentacoes(df, mes_ref, mes, ano)

#CAD E BENEFICIOS
beneficios      = cont_beneficio(df, mes_ref)


#TRABALHO
trabalho        = cont_trabalho(df, mes_ref)

#ESTRANGEIROS
estrangeiros    = cont_estrangeiros(df, mes_ref)

#JURIDICO
juridico        = cont_juridico(df, mes_ref)

#GERAR TABELA TAXA DE OCUPAÇÃO
df_dias = criar_tx_ocupacao(mes, ano)
df_dias = preencher_tx_ocupacao(df, mes_ref, df_dias, ini_acolhidos)
df_dias.to_excel(nome_tb_tx_ocupacao, index=False)

#VARIAVEIS PARA MROSC E DEVOLUTIVA

pvtn                    = movimentacoes["pvtn"]
comunitario             = movimentacoes["comunitario"]
familiar                = movimentacoes["familiar"]
ref_creas               = movimentacoes["ref_creas"]
transferidos            = movimentacoes["transferidos"]
desligados              = movimentacoes["desligados"]
total_acolhidos         = movimentacoes["total_acolhidos"]
obitos                  = movimentacoes["obitos"]
total_acolhidos_final   = movimentacoes["total_acolh_final"]
recepcionados           = movimentacoes["recepcionados"]


bf                      = beneficios["bf"]
bpc                     = beneficios["bpc"]
n_benef                 = beneficios["n_benef"]
outro_benef             = beneficios["outro_benef"]
inscr_cad               = beneficios["inscr_cad"]


formal                  = trabalho["formal"]
informal                = trabalho["informal"]


estrangeiros_cont       =  estrangeiros['estrangeiros']
refugiados              =  estrangeiros['refugiados']
imigrantes              =  estrangeiros['imigrantes']

cont_processo           = juridico["processo"]
cont_lib_condicional    = juridico["lib_condicional"]
cont_egressos           = juridico["egressos"]

#GERAR DEVOLUTIVA
df_devolutiva = gerar_devolutiva()
df_devolutiva = preencher_devolutiva(   df_devolutiva,
                                        pvtn,
                                        comunitario,
                                        familiar,
                                        ref_creas,
                                        transferidos,
                                        desligados,
                                        total_acolhidos,
                                        obitos,
                                        total_acolhidos_final,
                                        recepcionados,
                                        bf,
                                        bpc,
                                        n_benef,
                                        outro_benef,
                                        inscr_cad,
                                        formal,
                                        informal,
                                        estrangeiros_cont,
                                        refugiados,
                                        imigrantes,
                                        cont_processo,
                                        cont_lib_condicional,
                                        cont_egressos
                                     )
df_devolutiva.to_excel(nome_tb_devolutiva, index=False)

#GERAR MROSC
df_mrosc = gerar_mrosc()
df_mrosc = preencher_mrosc( df_mrosc,
                                        pvtn,
                                        comunitario,
                                        familiar,
                                        ref_creas,
                                        transferidos,
                                        desligados,
                                        total_acolhidos,
                                        obitos,
                                        total_acolhidos_final,
                                        recepcionados,
                                        bf,
                                        bpc,
                                        n_benef,
                                        outro_benef,
                                        inscr_cad,
                                        formal,
                                        informal,
                                        estrangeiros_cont,
                                        refugiados,
                                        imigrantes,
                                        cont_processo,
                                        cont_lib_condicional,
                                        cont_egressos
                                     )
df_mrosc.to_excel(nome_tb_mrosc, index=False)

print(df_mrosc)


#FORMATAÇÃO TABELA TAXA DE OCUPAÇÃO***********************************************************************
workbook    = load_workbook(nome_tb_tx_ocupacao)
sheet       = workbook.active

alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.alignment = alignment

# Ajusta o tamanho das colunas para melhor visualização
for col in sheet.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Obtém a letra da coluna
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = 13
    sheet.column_dimensions[col_letter].width = adjusted_width

workbook.save(nome_tb_tx_ocupacao)


#FORMATAÇÃO DA TABELA MROSC***************************************************************************
workbook = load_workbook(nome_tb_mrosc)
sheet    = workbook.active


#DIMENSIONAR LARGURA DAS COLUNAS 
sheet.column_dimensions['A'].width = 90
sheet.column_dimensions['B'].width = 15


#CENTRALIZAR CELULAS
for linha in range(1, 50):  # Suponha que 1000 linhas são suficientes
    sheet[f"B{linha}"].alignment = Alignment(horizontal='center', vertical='center')


#MESCLAR CELULAS 
sheet.merge_cells('A12:B12')
sheet.merge_cells('A16:B16')
sheet.merge_cells('A21:B21')
sheet.merge_cells('A29:B29')
sheet.merge_cells('A33:B33')
sheet.merge_cells('A41:B41')


#CENTRALIZAR CELULAS MESCLADAS
sheet['A12'].alignment = Alignment(horizontal='center', vertical='center')
sheet['A21'].alignment = Alignment(horizontal='center', vertical='center')
sheet['A29'].alignment = Alignment(horizontal='center', vertical='center')
sheet['A33'].alignment = Alignment(horizontal='center', vertical='center')
sheet['A16'].alignment = Alignment(horizontal='center', vertical='center')
sheet['A41'].alignment = Alignment(horizontal='center', vertical='center')

#QUEBRAR TEXTO AO TAMANHO DA COLUNA
sheet['A20'].alignment = Alignment(wrap_text=True)
sheet['A40'].alignment = Alignment(wrap_text=True)

#MUDAR COR DAS CELULAS
fill    = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
fill2   = PatternFill(start_color='5e5e5e', end_color='5e5e5e',fill_type='solid')

sheet['A12'].fill = fill
sheet['A16'].fill = fill  
sheet['A21'].fill = fill
sheet['A29'].fill = fill
sheet['A33'].fill = fill
sheet['A41'].fill = fill

#SALVAR TABELA
workbook.save(nome_tb_mrosc)

#FORMATAÇÃO TABELA DEVOLUTIVA**********************************************************************

workbook    = load_workbook(nome_tb_devolutiva)
sheet       = workbook.active

#MESCLAR CELULAS
sheet.merge_cells('A2:B2')
sheet.merge_cells('A5:B5')
sheet.merge_cells('A12:B12')
sheet.merge_cells('A14:B14')
sheet.merge_cells('A20:B20')
sheet.merge_cells('A22:B22')
sheet.merge_cells('A28:B28')
sheet.merge_cells('A32:B32')
sheet.merge_cells('A35:B35')


#CENTRALIZAR CELULAS MESCLADAS


#DIMENSIONAR LARGURA DAS COLUNAS 
sheet.column_dimensions['A'].width = 80
sheet.column_dimensions['B'].width = 15


#QUEBRAR TEXTO AO TAMANHO DA COLUNA
sheet['A21'].alignment = Alignment(wrap_text=True)
sheet['A37'].alignment = Alignment(wrap_text=True)


#CENTRALIZAR CELULAS
for linha in range(1, 50):  # Suponha que 1000 linhas são suficientes
    sheet[f"B{linha}"].alignment = Alignment(horizontal='center', vertical='center')


#MUDAR COR DAS CELULAS
sheet['A1'].fill    = fill2
sheet['B1'].fill    = fill2
sheet['A2'].fill    = fill
sheet['A5'].fill    = fill
sheet['A12'].fill   = fill
sheet['A14'].fill   = fill
sheet['A20'].fill   = fill
sheet['A22'].fill   = fill
sheet['A28'].fill   = fill
sheet['A32'].fill   = fill
sheet['A35'].fill   = fill

workbook.save(nome_tb_devolutiva)
